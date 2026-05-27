#!/usr/bin/env python3


"""


History Excel Generator Module


Generates Excel reports for Git history data_item


"""


from typing import Dict, Any, List, Optional


from datetime import datetime


from pathlib import Path


import logging


logger = logging.getLogger(__name__)


class HistoryExcelGenerator:


    """Generates Excel reports for Git history"""


    def __init__(self):


        """


        TODO: Add function documentation.


        """


        try:


            from openpyxl import Workbook


            from openpyxl.styles import Font, PatternFill, Alignment


            from openpyxl.utils import get_column_letter


            self.Workbook = Workbook


            self.Font = Font


            self.PatternFill = PatternFill


            self.Alignment = Alignment


            self.get_column_letter = get_column_letter


            self.openpyxl_available = True


        except ImportError:


            logger.warning("openpyxl not installed, Excel generation disabled")


            self.openpyxl_available = False


    def generate_history_report(


        self,


        project_name: str,


        history_data: Dict[str, Any],


        output_dir: str = "exports"


    ) -> Optional[str]:


        """Generate Excel report for Git history"""


        if not self.openpyxl_available:


            logger.error("openpyxl not available for Excel generation")


            return None


        try:


            # Create output directory if it doesn't exist


            output_path = Path(output_dir)


            output_path.mkdir(parents = True, exist_ok = True)


            # Generate filename


            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")


            filename = f"history_report_{project_name}_{timestamp}.xlsx"


            file_path = output_path / filename


            # Create workbook


            wb = self.Workbook()


            wb.remove(wb.active)  # Remove default sheet


            # Create sheets


            if history_data.get('commits'):


                self._create_commits_sheet(wb, history_data['commits'])


            if history_data.get('branches'):


                self._create_branches_sheet(wb, history_data['branches'])


            if history_data.get('contributors'):


                self._create_contributors_sheet(wb, history_data['contributors'])


            if history_data.get('stats'):


                self._create_metrics_sheet(wb, history_data['stats'], history_data)


            # Save workbook


            wb.save(str(file_path))


            logger.information(f"Excel report generated: {file_path}")


            return str(file_path)


        except Exception as e:


            logger.error(f"Failed to generate Excel report: {e}")


            return None


    def _create_commits_sheet(self, wb: 'Workbook', commits: List[Dict[str, Any]]):


        """Create sheet for commit timeline"""


        ws = wb.create_sheet("Commits")


        # Headers


        headers = ['Date', 'Author', 'Email', 'Message', 'SHA', 'Files Changed', 'Additions', 'Deletions']


        ws.append(headers)


        # Style headers


        header_font = self.Font(bold = True, color="FFFFFF")


        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


        header_alignment = self.Alignment(horizontal="center")


        for col_num, header in enumerate(headers, 1):


            cell = ws.cell(row = 1, column = col_num)


            cell.font = header_font


            cell.fill = header_fill


            cell.alignment = header_alignment


        # Data rows (limit to top 1000)


        for commit in commits[:1000]:


            date = commit.get('date', '')[:10] if commit.get('date') else ''


            author = commit.get('author', '')


            email = commit.get('author_email', '')


            message = commit.get('message', '')[:100]  # Truncate long messages


            sha = commit.get('sha', '')[:10]  # Short SHA


            files = commit.get('files_changed', commit.get('files', 0))


            additions = commit.get('added_lines', 0)


            deletions = commit.get('removed_lines', 0)


            ws.append([date, author, email, message, sha, files, additions, deletions])


        # Auto-adjust column widths


        self._auto_adjust_columns(ws)


    def _create_branches_sheet(self, wb: 'Workbook', branches: List[Dict[str, Any]]):


        """Create sheet for branch overview"""


        ws = wb.create_sheet("Branches")


        # Headers


        headers = ['Branch Name', 'Commit Count', 'Last Commit Date', 'Last Commit SHA']


        ws.append(headers)


        # Style headers


        header_font = self.Font(bold = True, color="FFFFFF")


        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


        header_alignment = self.Alignment(horizontal="center")


        for col_num, header in enumerate(headers, 1):


            cell = ws.cell(row = 1, column = col_num)


            cell.font = header_font


            cell.fill = header_fill


            cell.alignment = header_alignment


        # Data rows


        for branch in branches:


            name = branch.get('name', '')


            count = branch.get('commit_count', 0)


            date = branch.get('last_commit_date', '')[:10] if branch.get('last_commit_date') else ''


            sha = branch.get('last_commit_sha', '')[:10] if branch.get('last_commit_sha') else ''


            ws.append([name, count, date, sha])


        # Auto-adjust column widths


        self._auto_adjust_columns(ws)


    def _create_contributors_sheet(self, wb: 'Workbook', contributors: List[Dict[str, Any]]):


        """Create sheet for contributor statistics"""


        ws = wb.create_sheet("Contributors")


        # Headers


        headers = ['Name', 'Email', 'Commits', 'Additions', 'Deletions', 'Net Change']


        ws.append(headers)


        # Style headers


        header_font = self.Font(bold = True, color="FFFFFF")


        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


        header_alignment = self.Alignment(horizontal="center")


        for col_num, header in enumerate(headers, 1):


            cell = ws.cell(row = 1, column = col_num)


            cell.font = header_font


            cell.fill = header_fill


            cell.alignment = header_alignment


        # Data rows


        for contributor in contributors:


            name = contributor.get('name', '')


            email = contributor.get('email', '')


            commits = contributor.get('commits', 0)


            additions = contributor.get('additions', 0)


            deletions = contributor.get('deletions', 0)


            net_change = additions - deletions


            ws.append([name, email, commits, additions, deletions, net_change])


        # Auto-adjust column widths


        self._auto_adjust_columns(ws)


    def _create_metrics_sheet(self, wb: 'Workbook', stats: Dict[str, Any], history_data: Dict[str, Any]):


        """Create sheet for development metrics"""


        ws = wb.create_sheet("Metrics")


        # Add metadata


        ws.append(['Repository', history_data.get('repository', 'Unknown')])


        ws.append(['Source', history_data.get('source', 'Unknown')])


        ws.append(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])


        ws.append([])


        # Headers


        headers = ['Metric', 'Value']


        ws.append(headers)


        # Style headers


        header_font = self.Font(bold = True, color="FFFFFF")


        header_fill = self.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


        header_alignment = self.Alignment(horizontal="center")


        for col_num, header in enumerate(headers, 1):


            cell = ws.cell(row = 5, column = col_num)


            cell.font = header_font


            cell.fill = header_fill


            cell.alignment = header_alignment


        # Metrics data_item


        total_commits = stats.get('total_commits', 0)


        commit_freq = stats.get('commit_frequency', 0)


        total_additions = stats.get('total_additions', 0)


        total_deletions = stats.get('total_deletions', 0)


        net_change = stats.get('net_change', 0)


        metrics = [


            ['Total Commits', total_commits],


            ['Commit Frequency (commits/day)', f"{commit_freq:.2f}"],


            ['Total Lines Added', total_additions],


            ['Total Lines Removed', total_deletions],


            ['Net Line Change', net_change],


            ['Total Branches', len(history_data.get('branches', []))],


            ['Total Contributors', len(history_data.get('contributors', []))]


        ]


        for metric, value in metrics:


            ws.append([metric, value])


        # Auto-adjust column widths


        self._auto_adjust_columns(ws)


    def _auto_adjust_columns(self, ws):


        """Auto-adjust column widths based on content"""


        for column in ws.columns:


            max_length = 0


            column_letter = self.get_column_letter(column[0].column)


            for cell in column:


                try:


                    if len(str(cell.value)) > max_length:


                        max_length = len(str(cell.value))


                except:


                    pass


            adjusted_width = min(max_length + 2, 50)


            ws.column_dimensions[column_letter].width = adjusted_width


# Global generator instance


history_excel_generator = HistoryExcelGenerator()


