#!/usr/bin/env python3


"""


Excel Export Module


Exports analysis results to Excel format using openpyxl


"""


import os


from typing import Dict, Any, List, Optional


from datetime import datetime


from pathlib import Path


import logging


logger = logging.getLogger(__name__)


# Try to import openpyxl


try:


    from openpyxl import Workbook


    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


    from openpyxl.utils import get_column_letter


    OPENPYXL_AVAILABLE = True


except ImportError:


    OPENPYXL_AVAILABLE = False


    logger.warning("openpyxl not installed. Excel export will be disabled.")


class ExcelExportGenerator:


    """Generates Excel exports for analysis results"""


    def __init__(self):


        """Initialize Excel export generator"""


        self.enabled = OPENPYXL_AVAILABLE


    def export_analysis_to_excel(


        self,


        project_name: str,


        analysis_results: Dict[str, Any],


        output_path: Optional[str] = None


    ) -> Optional[str]:


        """Export analysis results to Excel file"""


        if not self.enabled:


            logger.warning("Excel export not enabled (openpyxl not installed)")


            return None


        if output_path is None:


            output_path = f"reports/{project_name}_analysis_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


        try:


            # Create output directory if it doesn't exist


            Path(output_path).parent.mkdir(parents = True, exist_ok = True)


            # Create workbook


            wb = Workbook()


            # Remove default sheet


            wb.remove(wb.active)


            # Add sheets


            summary_sheet = wb.create_sheet("Summary")


            code_structure_sheet = wb.create_sheet("Code Structure")


            code_quality_sheet = wb.create_sheet("Code Quality")


            security_sheet = wb.create_sheet("Security")


            technical_debt_sheet = wb.create_sheet("Technical Debt")


            performance_sheet = wb.create_sheet("Performance")


            recommendations_sheet = wb.create_sheet("Recommendations")


            # Populate sheets


            self._populate_summary_sheet(summary_sheet, project_name, analysis_results)


            self._populate_code_structure_sheet(code_structure_sheet, analysis_results)


            self._populate_code_quality_sheet(code_quality_sheet, analysis_results)


            self._populate_security_sheet(security_sheet, analysis_results)


            self._populate_technical_debt_sheet(technical_debt_sheet, analysis_results)


            self._populate_performance_sheet(performance_sheet, analysis_results)


            self._populate_recommendations_sheet(recommendations_sheet, analysis_results)


            # Set Summary as active sheet


            wb.active = summary_sheet


            # Save workbook


            wb.save(output_path)


            logger.information(f"Excel export generated: {output_path}")


            return output_path


        except Exception as e:


            logger.error(f"Failed to generate Excel export: {e}")


            return None


    def _populate_summary_sheet(self, sheet, project_name: str, data_item: Dict):


        """Populate summary sheet"""


        # Headers


        headers = ['Metric', 'Value']


        sheet.append(headers)


        # Style headers


        self._apply_header_style(sheet, 1)


        # Data


        metrics = [


            ['Project Name', project_name],


            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],


            ['Overall Score', data_item.get('overallScore', 'N/A')],


            ['Total Files', data_item.get('code_structure', {}).get('totalFiles', 'N/A')],


            ['Total Lines of Code', data_item.get('code_structure', {}).get('totalLines', 'N/A')],


            ['Code Quality Score', f"{data_item.get('code_quality', {}).get('codeQuality', 0)}%"],


            ['Security Score', f"{data_item.get('securityScore', 0)}/100"],


            ['Technical Debt Hours', data_item.get('totalHours', 'N/A')],


            ['Performance Score', f"{data_item.get('overallScore', 0)}/100"]


        ]


        for metric in metrics:


            sheet.append(metric)


        # Auto-adjust column widths


        sheet.column_dimensions['A'].width = 25


        sheet.column_dimensions['B'].width = 20


    def _populate_code_structure_sheet(self, sheet, data_item: Dict):


        """Populate code structure sheet"""


        code_structure = data_item.get('code_structure', {})


        # Headers


        headers = ['Metric', 'Value']


        sheet.append(headers)


        self._apply_header_style(sheet, 1)


        # Data


        metrics = [


            ['Total Files', code_structure.get('totalFiles', 0)],


            ['Total Lines of Code', code_structure.get('totalLines', 0)],


            ['Languages', ', '.join(code_structure.get('languages', []))],


            ['Architecture Pattern', code_structure.get('architecture', 'Unknown')],


            ['Main Patterns', ', '.join(code_structure.get('patterns', []))],


            ['Complexity Score', code_structure.get('complexity', 0)]


        ]


        for metric in metrics:


            sheet.append(metric)


        # Add file breakdown if available


        if 'files' in code_structure:


            sheet.append([])


            sheet.append(['File Breakdown'])


            sheet.append(['File', 'Lines', 'Language'])


            self._apply_header_style(sheet, sheet.max_row - 2)


            for file_info in code_structure['files'][:100]:  # Limit to 100 files


                sheet.append([


                    file_info.get('name', ''),


                    file_info.get('lines', 0),


                    file_info.get('language', '')


                ])


        sheet.column_dimensions['A'].width = 40


        sheet.column_dimensions['B'].width = 15


        sheet.column_dimensions['C'].width = 15


    def _populate_code_quality_sheet(self, sheet, data_item: Dict):


        """Populate code quality sheet"""


        code_quality = data_item.get('code_quality', {})


        # Headers


        headers = ['Metric', 'Value']


        sheet.append(headers)


        self._apply_header_style(sheet, 1)


        # Data


        metrics = [


            ['Code Quality Score', f"{code_quality.get('codeQuality', 0)}%"],


            ['Test Coverage', f"{code_quality.get('testCoverage', 0)}%"],


            ['Documentation Coverage', f"{code_quality.get('documentation', 0)}%"],


            ['Code Duplication', f"{code_quality.get('duplication', 0)}%"],


            ['Maintainability Index', code_quality.get('maintainability', 0)],


            ['Security Issues', code_quality.get('security_issues', 0)]


        ]


        for metric in metrics:


            sheet.append(metric)


        sheet.column_dimensions['A'].width = 30


        sheet.column_dimensions['B'].width = 15


    def _populate_security_sheet(self, sheet, data_item: Dict):


        """Populate security sheet"""


        # Headers


        headers = ['Metric', 'Value']


        sheet.append(headers)


        self._apply_header_style(sheet, 1)


        # Summary metrics


        metrics = [


            ['Security Score', f"{data_item.get('securityScore', 0)}/100"],


            ['Total Vulnerabilities', data_item.get('totalVulnerabilities', 0)],


            ['Critical Issues', data_item.get('criticalIssues', 0)],


            ['High Severity Issues', data_item.get('highSeverityIssues', 0)],


            ['Medium Severity Issues', data_item.get('mediumSeverityIssues', 0)],


            ['Low Severity Issues', data_item.get('lowSeverityIssues', 0)]


        ]


        for metric in metrics:


            sheet.append(metric)


        # Add vulnerability details if available


        vulnerabilities = data_item.get('dependencyVulnerabilities', [])


        if vulnerabilities:


            sheet.append([])


            sheet.append(['Vulnerability Details'])


            sheet.append(['Title', 'Severity', 'Package', 'CVE'])


            self._apply_header_style(sheet, sheet.max_row - 3)


            for vuln in vulnerabilities[:100]:  # Limit to 100


                sheet.append([


                    vuln.get('title', ''),


                    vuln.get('severity', ''),


                    vuln.get('package', ''),


                    vuln.get('id', '')


                ])


        sheet.column_dimensions['A'].width = 40


        sheet.column_dimensions['B'].width = 15


        sheet.column_dimensions['C'].width = 20


        sheet.column_dimensions['D'].width = 15


    def _populate_technical_debt_sheet(self, sheet, data_item: Dict):


        """Populate technical debt sheet"""


        # Headers


        headers = ['Metric', 'Value']


        sheet.append(headers)


        self._apply_header_style(sheet, 1)


        # Data


        metrics = [


            ['Total Debt Hours', data_item.get('totalHours', 0)],


            ['Debt Level', data_item.get('level', 'Unknown')],


            ['Estimated Cost', f"${data_item.get('estimatedCost', 0):,.2f}"],


            ['Priority', data_item.get('priority', 'Unknown')],


            ['Code Smell Debt Hours', data_item.get('smellDebtHours', 0)]


        ]


        for metric in metrics:


            sheet.append(metric)


        # Add code smell details if available


        code_smells = data_item.get('codeSmells', {}).get('smells', {})


        if code_smells:


            sheet.append([])


            sheet.append(['Code Smell Breakdown'])


            sheet.append(['Type', 'Count'])


            self._apply_header_style(sheet, sheet.max_row - 2)


            for smell_type, smell_list in code_smells.items():


                sheet.append([smell_type, len(smell_list)])


        sheet.column_dimensions['A'].width = 30


        sheet.column_dimensions['B'].width = 15


    def _populate_performance_sheet(self, sheet, data_item: Dict):


        """Populate performance sheet"""


        system_metrics = data_item.get('systemMetrics', {})


        # Headers


        headers = ['Metric', 'Value']


        sheet.append(headers)


        self._apply_header_style(sheet, 1)


        # Data


        cpu = system_metrics.get('cpu', {})


        memory = system_metrics.get('memory', {})


        metrics = [


            ['Overall Score', f"{data_item.get('overallScore', 0)}/100"],


            ['Uptime (seconds)', f"{data_item.get('uptime', 0):.0f}"],


            ['CPU Usage', f"{cpu.get('current', 0)}%"],


            ['CPU Average', f"{cpu.get('average', 0)}%"],


            ['CPU Status', cpu.get('status', 'Unknown')],


            ['Memory Usage', f"{memory.get('current', 0)}%"],


            ['Memory Available (GB)', f"{memory.get('available_gb', 0):.2f}"],


            ['Memory Used (GB)', f"{memory.get('used_gb', 0):.2f}"],


            ['Memory Status', memory.get('status', 'Unknown')]


        ]


        for metric in metrics:


            sheet.append(metric)


        sheet.column_dimensions['A'].width = 30


        sheet.column_dimensions['B'].width = 20


    def _populate_recommendations_sheet(self, sheet, data_item: Dict):


        """Populate recommendations sheet"""


        recommendations = data_item.get('recommendations', [])


        if not recommendations:


            sheet.append(['No recommendations available'])


            return


        # Headers


        headers = ['#', 'Priority', 'Type', 'Message', 'Action']


        sheet.append(headers)


        self._apply_header_style(sheet, 1)


        # Data


        for i, rec in enumerate(recommendations, 1):


            sheet.append([


                i,


                rec.get('priority', 'medium'),


                rec.get('type', 'general'),


                rec.get('message', ''),


                rec.get('action', '')


            ])


        sheet.column_dimensions['A'].width = 5


        sheet.column_dimensions['B'].width = 12


        sheet.column_dimensions['C'].width = 15


        sheet.column_dimensions['D'].width = 50


        sheet.column_dimensions['E'].width = 40


    def _apply_header_style(self, sheet, row_num: int):


        """Apply header styling to a row"""


        header_font = Font(bold = True, color='FFFFFF')


        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')


        header_alignment = Alignment(horizontal='center', vertical='center')


        thin_border = Border(


            left = Side(style='thin'),


            right = Side(style='thin'),


            top = Side(style='thin'),


            bottom = Side(style='thin')


        )


        for cell in sheet[row_num]:


            cell.font = header_font


            cell.fill = header_fill


            cell.alignment = header_alignment


            cell.border = thin_border


# Global Excel export generator instance


excel_export = ExcelExportGenerator()


